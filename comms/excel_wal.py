# comms/excel_wal.py
# Write-Ahead Log (WAL) + Excel logger.
#
# Architecture:
#   1. Every scan appended to CSV WAL immediately (near-zero latency, durable).
#   2. Excel built ONCE on close_batch() — never during logging.
#      This eliminates the O(n) scan-thread stall that caused slowdowns
#      after ~5 000 scans when Excel was rebuilt every 5 records.
#
# File layout (per device):
#   logs/Device1/ProductionLog_<batch>_<timestamp>.xlsx
#   logs/Device1/wal/WAL_<batch>_<timestamp>.csv

import os
import csv
import logging
from datetime import datetime
from typing import Optional, Callable

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill

from core.models import ReadRecord, FILE_PREFIX, WAL_PREFIX

log = logging.getLogger("pharma.excel_wal")

WAL_FIELDS = [
    "read_id", "timestamp", "batch_id", "operator_id",
    "product_name", "raw_data", "master_data", "status",
]

EXCEL_COLUMNS = [
    ("Read ID",     10),
    ("Time",        22),
    ("Pharma Code", 20),
    ("Status",      10),
]

PASS_FILL   = PatternFill("solid", fgColor="C6EFCE")
PASS_FONT   = Font(color="276221", bold=True)
FAIL_FILL   = PatternFill("solid", fgColor="FFC7CE")
FAIL_FONT   = Font(color="9C0006", bold=True)
HEADER_FONT = Font(bold=True)


class WALExcelLogger:

    def __init__(self, log_dir: str, wal_dir: str):
        self.log_dir     = log_dir
        self.wal_dir     = wal_dir

        self._excel_path:   Optional[str]           = None
        self._wal_path:     Optional[str]            = None
        self._wal_file:     Optional[object]         = None
        self._wal_writer:   Optional[csv.DictWriter] = None
        self._batch_id:     str                      = ""
        self._operator_id:  str                      = ""
        self._product_name: str                      = ""
        self._started:      Optional[datetime]       = None

        # CFR21: set by AppController before start() — used for integrity seal
        self.cfr_user  = None   # cfr21.user_manager.User or None
        self.device_id = 0      # 1 or 2
        self.regulated_records = None
        self.regulated_batch_id: str = ""

    # ── path helpers ──────────────────────────────────────────────────────────

    def _make_excel_path(self, batch_id: str, started: datetime) -> str:
        safe  = _safe_name(batch_id)
        fname = f"{FILE_PREFIX}_{safe}_{started.strftime('%Y%m%d_%H%M%S')}.xlsx"
        return os.path.join(self.log_dir, fname)

    def _make_wal_path(self, batch_id: str, started: datetime) -> str:
        safe  = _safe_name(batch_id)
        fname = f"{WAL_PREFIX}_{safe}_{started.strftime('%Y%m%d_%H%M%S')}.csv"
        return os.path.join(self.wal_dir, fname)

    # ── batch lifecycle ───────────────────────────────────────────────────────

    def start_batch(self, batch_id: str, started: datetime) -> tuple[str, str]:
        """Open WAL for append. Returns (excel_path, wal_path)."""
        os.makedirs(self.log_dir, exist_ok=True)
        os.makedirs(self.wal_dir, exist_ok=True)

        self._batch_id    = batch_id
        self._started     = started
        self._excel_path  = self._make_excel_path(batch_id, started)
        self._wal_path    = self._make_wal_path(batch_id, started)

        if os.path.exists(self._wal_path):
            log.warning("WAL exists — crash recovery mode: %s", self._wal_path)
            # Restore operator/product from existing WAL records
            self._restore_meta_from_wal()

        if self.regulated_records is None or not self.regulated_batch_id:
            self._wal_file = open(self._wal_path, "a", newline="", encoding="utf-8")
            self._wal_writer = csv.DictWriter(self._wal_file, fieldnames=WAL_FIELDS)
            if self._wal_file.tell() == 0:
                self._wal_writer.writeheader()
                self._wal_file.flush()

        log.info("Batch started | Excel: %s | WAL: %s",
                 self._excel_path, self._wal_path)
        return self._excel_path, self._wal_path

    def append_record(self, rec: ReadRecord):
        """
        Append one scan record to the WAL CSV.
        Fast — single row write + flush. No Excel work here.
        """
        if self.regulated_records is not None and self.regulated_batch_id:
            self._operator_id = rec.operator_id
            self._product_name = rec.product_name
            return
        if self._wal_writer is None:
            log.error("append_record called before start_batch")
            return

        if not self._operator_id:
            self._operator_id  = rec.operator_id
        if not self._product_name:
            self._product_name = rec.product_name

        row = {
            "read_id":      rec.read_id,
            "timestamp":    rec.timestamp.strftime("%Y-%m-%d %H:%M:%S %Z"),
            "batch_id":     rec.batch_id,
            "operator_id":  rec.operator_id,
            "product_name": rec.product_name,
            "raw_data":     rec.raw_data,
            "master_data":  rec.master_data,
            "status":       rec.status,
        }
        self._wal_writer.writerow(row)
        self._wal_file.flush()
        log.debug("WAL appended: read_id=%s status=%s", rec.read_id, rec.status)

    def close_batch(self,
                    progress_callback: Optional[Callable[[int, int], None]] = None):
        """
        Close the batch:
          1. Flush and close WAL
          2. Build Excel from WAL (once — with progress reporting)
          3. Seal SHA-256 checksums into compliance.db

        progress_callback(current, total) is called during Excel build so
        the GUI can update a progress bar without polling.
        """
        log.info("Closing batch — building Excel from WAL")

        # Close WAL first so Excel reads a complete, consistent file
        self._close_wal()

        if self.regulated_records is not None and self.regulated_batch_id:
            self._write_wal_from_authoritative_records()

        # Build Excel with progress reporting
        self._rebuild_excel_from_wal(progress_callback=progress_callback)

        # CFR21: seal file checksums into compliance.db
        try:
            from cfr21.record_integrity import seal_batch_files
            seal_batch_files(
                user       = self.cfr_user,
                batch_id   = self._batch_id,
                device_id  = self.device_id,
                excel_path = self._excel_path,
                wal_path   = self._wal_path,
            )
        except Exception as e:
            log.error("Integrity seal failed (batch still closed): %s", e)

    def get_wal_counts(self) -> tuple[int, int]:
        """
        Read the WAL and return (pass_count, fail_count).
        Used for counter restore when the same batch ID is re-entered
        after a crash or emergency shutdown.
        Returns (0, 0) if WAL does not exist or cannot be read.
        """
        if not self._wal_path or not os.path.exists(self._wal_path):
            return 0, 0
        return self._count_wal(self._wal_path)

    @staticmethod
    def find_existing_wal(batch_id: str, wal_dir: str) -> Optional[str]:
        """
        Search wal_dir for an existing WAL file for the given batch_id.
        Returns the most recent matching file path, or None if not found.
        Used on startup/session-setup to detect crash recovery scenarios.
        """
        if not os.path.isdir(wal_dir):
            return None
        safe = _safe_name(batch_id)
        prefix = f"{WAL_PREFIX}_{safe}_"
        matches = [
            os.path.join(wal_dir, f)
            for f in os.listdir(wal_dir)
            if f.startswith(prefix) and f.endswith(".csv")
        ]
        if not matches:
            return None
        # Return most recently modified file
        return max(matches, key=os.path.getmtime)

    @staticmethod
    def _count_wal(wal_path: str) -> tuple[int, int]:
        """Count PASS and FAIL records in a WAL CSV file."""
        passes = fails = 0
        try:
            with open(wal_path, "r", encoding="utf-8", newline="") as f:
                for row in csv.DictReader(f):
                    if row.get("status") == "PASS":
                        passes += 1
                    else:
                        fails += 1
        except Exception as e:
            log.error("_count_wal failed for %s: %s", wal_path, e)
        return passes, fails

    # ── Excel rebuild ─────────────────────────────────────────────────────────

    def _rebuild_excel_from_wal(self,
                                 progress_callback: Optional[Callable[[int, int], None]] = None):
        if not self._wal_path or not os.path.exists(self._wal_path):
            log.warning("_rebuild_excel_from_wal: WAL not found — skipping")
            return
        try:
            records = self._read_wal()
        except Exception as e:
            log.error("Failed to read WAL for Excel rebuild: %s", e)
            return

        total = len(records)
        if progress_callback:
            progress_callback(0, total)

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Log"

            # Header metadata block
            ws["A1"] = f"Batch ID:      {self._batch_id}"
            ws["A2"] = f"Operator ID:   {self._operator_id or '–'}"
            ws["A3"] = f"Product Name:  {self._product_name or '–'}"
            ws["A4"] = f"Started:       {self._started.strftime('%Y-%m-%d %I:%M:%S %p')}"
            ws.append([])

            # Column headers
            headers = [col[0] for col in EXCEL_COLUMNS]
            ws.append(headers)
            header_row = ws.max_row
            for cell in ws[header_row]:
                cell.font = HEADER_FONT
            for i, (_, width) in enumerate(EXCEL_COLUMNS, start=1):
                ws.column_dimensions[get_column_letter(i)].width = width

            # Data rows
            for idx, r in enumerate(records, start=1):
                ws.append([r["read_id"], r["timestamp"],
                           r["raw_data"], r["status"]])
                row_idx = ws.max_row
                is_pass = r["status"] == "PASS"
                fill    = PASS_FILL if is_pass else FAIL_FILL
                sfont   = PASS_FONT if is_pass else FAIL_FONT
                for col in range(1, len(EXCEL_COLUMNS) + 1):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.fill = fill
                    if col == len(EXCEL_COLUMNS):
                        cell.font = sfont

                # Report progress every 100 rows to avoid hammering the GUI
                if progress_callback and idx % 100 == 0:
                    progress_callback(idx, total)

            wb.save(self._excel_path)

            if progress_callback:
                progress_callback(total, total)

            log.info("Excel built: %s rows → %s", total, self._excel_path)

        except Exception as e:
            log.error("Excel rebuild failed: %s", e)

    def _read_wal(self) -> list[dict]:
        records = []
        with open(self._wal_path, "r", encoding="utf-8", newline="") as f:
            for row in csv.DictReader(f):
                records.append(row)
        return records

    def _write_wal_from_authoritative_records(self):
        """Generate the compatibility WAL only from committed database scans."""
        if not self._wal_path:
            raise RuntimeError("No derived WAL path is configured.")
        _, records = self.regulated_records.get_batch_record(
            self._batch_id, self.device_id)
        with open(self._wal_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=WAL_FIELDS)
            writer.writeheader()
            for record in records:
                writer.writerow({
                    "read_id": record["sequence_no"],
                    "timestamp": record["recorded_at"],
                    "batch_id": self._batch_id,
                    "operator_id": record["operator_id"],
                    "product_name": record["product_name"],
                    "raw_data": record["raw_data"],
                    "master_data": record["master_data"],
                    "status": record["status"],
                })

    def _restore_meta_from_wal(self):
        """Read operator/product from first WAL record (crash recovery)."""
        try:
            with open(self._wal_path, "r", encoding="utf-8", newline="") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    self._operator_id  = row.get("operator_id", "")
                    self._product_name = row.get("product_name", "")
                    break
        except Exception as e:
            log.warning("Could not restore meta from WAL: %s", e)

    def _close_wal(self):
        if self._wal_file is not None:
            try:
                self._wal_file.close()
            except Exception as e:
                log.warning("WAL close error: %s", e)
            self._wal_file   = None
            self._wal_writer = None

    @property
    def excel_path(self) -> Optional[str]:
        return self._excel_path

    @property
    def wal_path(self) -> Optional[str]:
        return self._wal_path


def _safe_name(text: str) -> str:
    return "".join(c if c.isalnum() or c in "-_" else "_" for c in text)
